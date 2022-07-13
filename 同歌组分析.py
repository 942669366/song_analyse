import openpyxl
import time
import tkinter as tk
from tkinter.messagebox import showinfo
from tkinter import filedialog
import datetime


window = tk.Tk()
window.title('歌词信息提取')
window.geometry('500x250')
window.config(background = '#d6d6d6')


lie1 = 20
lie2 = 190

tk.Label(window, text="请选择类别表数据文件：",bg = "#d6d6d6").place(x=lie1, y=110)
var_name = tk.StringVar()  # 文件输入路径变量

entry_name = tk.Entry(window, textvariable=var_name, width=25)
entry_name.place(x=lie2, y=110)


def selectPath_file():#文件选择
    path_ = filedialog.askopenfilename(filetypes=[("数据表", [".xlsx"])])
    print(path_)
    var_name.set(path_)


"""
操作顺序：
1、取表中所有id,放入超音数<7、同歌组数据>接口一中获取数据，存入表中的Sheet2
2、vlookup()将跑取的月播放量和翻唱自id匹配到Sheet1中
3、获取接口一中跑出来的翻唱自id字段的数据,放入超音数接口二中获取数据，存入表中的Sheet3
4、执行程序
"""



music_list = []
#同歌组去重
def max_group_id_s(sheet,row_num,sheet2,row_num2):
    max_group_id_lsit = []
    for k in range(3,row_num+1):
        max_group_id = sheet.cell(k, 16).value
        if max_group_id not in max_group_id_lsit:
            max_group_id_lsit.append(max_group_id)
    return max_group_id_lsit

#同歌组播放量求和生成字典
def max_group_id_disa(sheet,row_num,sheet2,row_num2):
    max_group_id_list = max_group_id_s(sheet,row_num,sheet2,row_num2)
    zyqk_song_id_tme_3plat_js_play_cnt_dis_all = {}#判断播放量排名
    max_group_id_diss = {}
    zyqk_song_id_list_diss = {}
    jichuxinxi_diss = {}
    for j in max_group_id_list:
        tme_3plat_js_play_cnt_liss = []
        zyqk_song_id_tme_3plat_js_play_cnt = {}
        zyqk_song_id_list = []
        jichuxinxi_list2= []
        for i in range(3,row_num+1):
            zyqk_song_id = sheet.cell(i, 1).value
            song_name = sheet.cell(i, 2).value
            singer_name = sheet.cell(i, 5).value
            max_group_id = sheet.cell(i, 16).value
            tme_3plat_js_play_cnt = sheet.cell(i, 19).value
            jichuxinxi_diss2_1 = {}
            if j == max_group_id:
                '''
                基础信息字典
                '''
                jichuxinxi_diss2_1['zyqk_song_id'] = zyqk_song_id
                jichuxinxi_diss2_1['song_name'] = song_name
                jichuxinxi_diss2_1['max_group_id'] = max_group_id
                jichuxinxi_diss2_1['singer_name'] = singer_name
                jichuxinxi_list2.append(jichuxinxi_diss2_1)

                zyqk_song_id_tme_3plat_js_play_cnt[zyqk_song_id] = tme_3plat_js_play_cnt
                zyqk_song_id_list.append(zyqk_song_id)
                # 同歌组总播放量
                tme_3plat_js_play_cnt_liss.append(tme_3plat_js_play_cnt)

            else:
                continue
        zyqk_song_id_list_diss[j] = zyqk_song_id_list
        max_group_id_diss[j] = sum(tme_3plat_js_play_cnt_liss)#同歌组下歌曲总播放量
        zyqk_song_id_tme_3plat_js_play_cnt_dis_all[j] = zyqk_song_id_tme_3plat_js_play_cnt#同个组下歌曲id下对应的播放量
        jichuxinxi_diss[j] = jichuxinxi_list2#同个组下歌曲id对应的歌曲信息

    #max_group_id_diss 同歌组下歌曲总播放量
    #zyqk_song_id_tme_3plat_js_play_cnt_dis_all 同歌组下单个id对应的播放量
    return zyqk_song_id_tme_3plat_js_play_cnt_dis_all,max_group_id_diss,zyqk_song_id_list_diss,jichuxinxi_diss,max_group_id_list

#标杆原唱
def biaoganyuanchang(k_y_and_y_k_diss_all,jichuxinxi_diss):
    k_y_and_y_k_diss_all2 = {}
    for p in k_y_and_y_k_diss_all.keys():
        k_y_and_y_k_diss_all_values = k_y_and_y_k_diss_all[p]
        k_y_and_y_k_diss = {}
        for k_y_and_y_k_key, k_y_and_y_k_values in k_y_and_y_k_diss_all_values.items():
            if k_y_and_y_k_values == 1:
                k_y_and_y_k_diss["biaogan_yuancheng_id"] = k_y_and_y_k_key
            else:
                continue
        k_y_and_y_k_diss_all2[p] = k_y_and_y_k_diss

    for kl in k_y_and_y_k_diss_all2.keys():
        for zyqk_song_id2 in jichuxinxi_diss[kl]:
            if k_y_and_y_k_diss_all2[kl]['biaogan_yuancheng_id'] == zyqk_song_id2['zyqk_song_id']:
                k_y_and_y_k_diss_all2[kl]['song_name'] = zyqk_song_id2['song_name']
                k_y_and_y_k_diss_all2[kl]['singer_name'] = zyqk_song_id2['singer_name']
                k_y_and_y_k_diss_all2[kl]['max_group_id'] = zyqk_song_id2['max_group_id']
            else:
                continue

    return k_y_and_y_k_diss_all2

#检验id
def jianyan(max_group_id_list,sheet,row_num,sheet2,row_num2):
    jiaoyan_diss = {}
    for j in max_group_id_list:
        jiaoyan_list = []
        for i in range(3,row_num+1):
            max_group_id = sheet.cell(i, 16).value
            jianyanid = sheet.cell(row=i, column=31).value
            if j == max_group_id:
                if jianyanid == 0:
                    continue
                else:
                    jiaoyan_list.append(jianyanid)
            else:
                continue
        jiaoyan_diss[j] = jiaoyan_list

    for k,o in jiaoyan_diss.items():
        if len(o) == 0:
            jiaoyan_diss[k] = '无'
        else:
            jiaoyan_max = max(o, key=o.count)
            jiaoyan_diss[k] = jiaoyan_max

    return jiaoyan_diss

#校验id原唱
def jiaoyan_ids_yuanchang(jiaoyan_diss,sheet,row_num,sheet2,row_num2):
    jiaoyan_id_yuanchang_diss = {}
    for op1,op2 in jiaoyan_diss.items():
        for k1 in range(2, row_num2 + 1):
            idyuanchang_id = sheet2.cell(k1,1).value
            idyuanchang = sheet2.cell(k1,3).value
            if op2 == idyuanchang_id:
                jiaoyan_id_yuanchang_diss[op1] = idyuanchang
            else:
                continue
    return jiaoyan_id_yuanchang_diss

#如分别3个id满足3个指标且至少2位歌手相同则以相同歌手为原唱，播放最高为标杆。如歌手均不同或相同歌手非播放最高则返回运营审核
def zhibiao_fensan(sheet,row_num,sheet2,row_num2):
    max_group_id_lsit = max_group_id_s()
    idds = {}
    for a1 in max_group_id_lsit:
        bbb = {}
        for i in range(3, row_num + 1):
            zyqk_song_id1 = sheet.cell(i, 1).value  # 歌曲id
            max_group_id1 = sheet.cell(i, 16).value  # 同个组id
            bofangliangpaiming1 = sheet.cell(i, 28).value
            rukupaiming1 = sheet.cell(i, 29).value
            jiaquanpaimings1 = sheet.cell(i, 30).value
            if a1 == max_group_id1:
                if bofangliangpaiming1 == 1 and rukupaiming1 != 1 and jiaquanpaimings1 != 1:
                    bbb[zyqk_song_id1] = '播放量'
                elif bofangliangpaiming1 != 1 and rukupaiming1 == 1 and jiaquanpaimings1 != 1:
                    bbb[zyqk_song_id1] = '入库排名'
                elif bofangliangpaiming1 != 1 and rukupaiming1 != 1 and jiaquanpaimings1 == 1:
                    bbb[zyqk_song_id1] = '加权排名'
                else:
                    continue
            else:
                continue
        idds[a1] = bbb

    max_group_id_listaaa = []
    for mm, s in idds.items():
        if len(s) > 2:
            if '播放量' in s.values() and '入库排名' in s.values() and '加权排名' in s.values():
                if idds.keys() not in max_group_id_listaaa:
                    max_group_id_listaaa.append(mm)
                else:
                    continue
            else:
                continue
        else:
            continue

    max_group_id_listaaa2 = {}

    for nn in max_group_id_listaaa:
        max_group_id_listaaa2[nn] = idds[nn]

    return max_group_id_listaaa2

# k_y_all 同歌组播放量排名
# y_k_all 同歌组入库时间排名
def paiming(z_t_all):
    k_y_all = {}
    y_k_all = {}
    for keyy, valuess in z_t_all.items():
        valuess_list = sorted(valuess.items(), key=lambda kv: (kv[1], kv[0]))
        valuess_list.reverse()
        d_k = {}
        for valuesa in range(len(valuess_list)):
            d_k[valuess_list[valuesa][0]] = valuesa + 1
        k_y_all[keyy] = d_k

        keys_list = sorted(valuess.keys())
        keys_list.sort()
        k_d = {}

        for keyss in range(len(keys_list)):
            k_d[keys_list[keyss]] = keyss + 1
        y_k_all[keyy] = k_d

    # k_y_all 同歌组播放量排名
    # y_k_all 同歌组入库时间排名
    return k_y_all, y_k_all

#加权排名
def jiaquanpaiming(k_y_all,y_k_all,zyqk_song_id_list_diss):
    k_y_and_y_k_diss_all = {}
    for j in zyqk_song_id_list_diss.keys():
        k_y_and_y_k_diss = {}
        for k in zyqk_song_id_list_diss[j]:
            k_y_and_y_k = k_y_all[j][k]+y_k_all[j][k]
            k_y_and_y_k_diss[k] = k_y_and_y_k
        k_y_and_y_k_diss_all[j] = k_y_and_y_k_diss
    k_y_alls = {}
    for keyy, valuess in k_y_and_y_k_diss_all.items():
        valuessa_list = sorted(valuess.items(), key=lambda kv: (kv[1], kv[0]))
        valuessa_list.reverse()
        valuessa_list.reverse()
        d_kl = {}
        for valuesa in range(len(valuessa_list)):
            d_kl[valuessa_list[valuesa][0]] = valuesa + 1
        k_y_alls[keyy] = d_kl

    return k_y_alls

#返回列表中出现最多的元素
def showmax(It):
    index1 = 0        #记录出现次数最多的元素下表
    max = 0             #记录最大的元素出现次数
    for i in range(len(It)):
        flag = 0        #记录每个元素出现的次数
        for j in range(i+1,len(It)):    #遍历i之后的元素下标
            if It[j] == It[i]:
                flag +=1                #每当发现与自己相同的元素 flag+1
        if flag > max:                  #如果此时元素出现的次数大于最大值，记录此时元素的下标
            max = flag+1
            index1 = i
    return It[index1],max   #返回出现最多的元素和重复的次数


def chuli(zyqk_song_id,song_name,singer_name,max_group_id,jiaoyan_id_yuanchang,max_group_id_diss,jiaoyan_diss,jiaoyan_id_yuanchang_diss,k_y_and_y_k_diss_all2,k_y_all,y_k_all,k_y_and_y_k_diss_all):
    tonggezu_all_data = max_group_id_diss[max_group_id]
    if zyqk_song_id == k_y_and_y_k_diss_all2[max_group_id][
        'biaogan_yuancheng_id'] and max_group_id in k_y_and_y_k_diss_all2.keys():
        biaogan = '标杆'
    else:
        biaogan = '非标杆'

    if max_group_id in k_y_and_y_k_diss_all2.keys() and zyqk_song_id in k_y_all[max_group_id].keys() and singer_name == \
            k_y_and_y_k_diss_all2[max_group_id]['singer_name'] and k_y_and_y_k_diss_all2[max_group_id][
        'song_name'] in song_name:
        yuanfanchang_jieguo = '原唱'
    else:
        yuanfanchang_jieguo = '翻唱'
    jianyanid = jiaoyan_diss[max_group_id]
    if max_group_id not in jiaoyan_id_yuanchang_diss:
        jiaoyan_id_yuanchang = '无'
    else:
        jiaoyan_id_yuanchang = jiaoyan_id_yuanchang_diss[max_group_id]
    zuyuanchang = k_y_and_y_k_diss_all2[max_group_id]['singer_name']


    if '|' in zuyuanchang and '|' in jiaoyan_id_yuanchang:
        if len([i for i in zuyuanchang.split('|') if i in jiaoyan_id_yuanchang.split('|')]) == len(
                zuyuanchang.split('|')) == len(jiaoyan_id_yuanchang.split('|')):
            shifoujianyan = '否'
        else:
            shifoujianyan = '是，原唱与校验不同'
    else:
        if zuyuanchang == jiaoyan_id_yuanchang:
            shifoujianyan = '否'
        else:
            shifoujianyan = '是，原唱与校验不同'
    bofangliangpaiming = y_k_all[max_group_id][zyqk_song_id]
    rukupaiming = k_y_all[max_group_id][zyqk_song_id]
    jiaquan = k_y_and_y_k_diss_all[max_group_id][zyqk_song_id]
    return tonggezu_all_data, biaogan, yuanfanchang_jieguo,bofangliangpaiming, rukupaiming, jiaquan,  jianyanid, jiaoyan_id_yuanchang,zuyuanchang, shifoujianyan


def mains():
    start = datetime.datetime.now()
    path = var_name.get()
    wb = openpyxl.load_workbook(path)
    sheet = wb['Sheet1']
    row_num = sheet.max_row
    sheet2 = wb['Sheet3']
    row_num2 = sheet2.max_row

    zyqk_song_id_tme_3plat_js_play_cnt_dis_all, max_group_id_diss,zyqk_song_id_list_diss,jichuxinxi_diss,max_group_id_list = max_group_id_disa(sheet,row_num,sheet2,row_num2)

    k_y_all, y_k_all = paiming(zyqk_song_id_tme_3plat_js_play_cnt_dis_all)

    k_y_and_y_k_diss_all = jiaquanpaiming(k_y_all, y_k_all,zyqk_song_id_list_diss)

    k_y_and_y_k_diss_all2 = biaoganyuanchang(k_y_and_y_k_diss_all,jichuxinxi_diss)#标杆原唱信息

    jiaoyan_diss = jianyan(max_group_id_list,sheet,row_num,sheet2,row_num2)
    jiaoyan_id_yuanchang_diss = jiaoyan_ids_yuanchang(jiaoyan_diss,sheet,row_num,sheet2,row_num2)

    for i in range(3, row_num + 1):
        zyqk_song_id = sheet.cell(i, 1).value#歌曲id
        song_name = sheet.cell(i, 2).value
        singer_name = sheet.cell(i, 5).value
        max_group_id = sheet.cell(i, 16).value#同个组id
        jiaoyan_id_yuanchang = sheet.cell(i, 25).value
        tonggezu_all_data, biaogan, yuanfanchang_jieguo,bofangliangpaiming, rukupaiming, jiaquan,  jianyanid, jiaoyan_id_yuanchang,zuyuanchang, shifoujianyan = chuli(zyqk_song_id,song_name,singer_name,max_group_id,jiaoyan_id_yuanchang,max_group_id_diss,jiaoyan_diss,jiaoyan_id_yuanchang_diss,k_y_and_y_k_diss_all2,k_y_all,y_k_all,k_y_and_y_k_diss_all)

        sheet.cell(row=i, column=20).value = tonggezu_all_data
        sheet.cell(row=i, column=21).value = biaogan
        sheet.cell(row=i, column=22).value = yuanfanchang_jieguo
        sheet.cell(row=i, column=23).value = jianyanid
        sheet.cell(row=i, column=24).value = zuyuanchang
        sheet.cell(row=i, column=25).value = jiaoyan_id_yuanchang
        sheet.cell(row=i, column=26).value = shifoujianyan
        sheet.cell(row=i, column=28).value = rukupaiming
        sheet.cell(row=i, column=29).value = bofangliangpaiming
        sheet.cell(row=i, column=30).value = jiaquan

        print(str(zyqk_song_id)+'--------------------' + str(max_group_id)+'--------------------'+str(tonggezu_all_data)+'--------------------'+str(biaogan)+'--------------------'+str(yuanfanchang_jieguo)+'--------------------'+str(jianyanid)+'--------------------'+str(zuyuanchang)+'--------------------'+str(jiaoyan_id_yuanchang)+'--------------------'+str(shifoujianyan)+'--------------------'+str(rukupaiming)+'--------------------'+str(bofangliangpaiming)+'--------------------'+str(jiaquan))
    wb.save(path)

    # music_list12 = []
    # for i_id in jiaoyan_diss.values():
    #     if i_id == '无':
    #         continue
    #     else:
    #         if str(i_id) not in music_list12:
    #             continue
    #         else:
    #             music_list12.append(str(i_id) + ',')
    # f = open('./文件/校验id.txt', 'w')
    # f.writelines(music_list12)
    # f.close()

# if __name__ == '__main__':
#     mains()

    end = datetime.datetime.now()
    showinfo('提示', '运行完毕！' + '\n' + '程序运行时间: ' + str(((end - start) / 60).seconds) + '分钟')
    window.quit()

tk.Button(window, text="文件选择", command=selectPath_file,bg = '#d1d1d1').place(x=400, y=105)
tk.Button(window, text="运行", command=mains,width = 12,height = 1,bg = '#d1d1d1').place(x=210, y=200)
window.mainloop()  # 显示窗口